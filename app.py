import os
import io
from datetime import datetime, date
from flask import (Flask, render_template, request, redirect,
                   url_for, session, jsonify, send_file)
from flask_wtf.csrf import CSRFProtect
import tracker as db

app = Flask(__name__)

# ── Security: secret key from environment, never hardcoded ────────────────────
app.secret_key = os.environ.get("SECRET_KEY") or os.urandom(32)
app.config["WTF_CSRF_ENABLED"] = os.environ.get("TESTING") != "1"

# Cookie hardening — Vercel serves over HTTPS, so require Secure cookies in
# production; SameSite=Lax stops most CSRF/XSS-adjacent cookie leakage while
# still allowing normal top-level navigation (e.g. following a login redirect).
app.config["SESSION_COOKIE_SECURE"]   = os.environ.get("TESTING") != "1"
app.config["SESSION_COOKIE_SAMESITE"] = "Lax"
app.config["SESSION_COOKIE_HTTPONLY"] = True

csrf = CSRFProtect(app)

db.seed_admin()

CATS = ["Food", "Transport", "Shopping", "Bills",
        "Health", "Entertainment", "Education", "Other"]

def base_ctx(user):
    """Variables every page needs to avoid Jinja undefined errors."""
    from datetime import date as _d
    today = _d.today()
    rows  = db.view_expenses_records(user)
    cat_t = {}
    for r in rows:
        cat_t[r[3]] = cat_t.get(r[3], 0) + r[2]
    # monthly trend: last 6 months totals
    trend = []
    for mo in range(5, -1, -1):
        import datetime
        d = today.replace(day=1)
        # go back mo months
        month = (today.month - mo - 1) % 12 + 1
        year  = today.year - ((mo + (12 - today.month)) // 12)
        mo_rows = db.get_expenses_for_month(user, month, year)
        trend.append(sum(r[2] for r in mo_rows))
    return dict(
        expenses=rows,          # was never passed before — template needs it on every page
        chart_data=cat_t,       # feeds the donut chart / report bars, previously undefined
        cat_totals=cat_t,
        monthly_trend=trend,
        budget_alerts=db.get_notifications(user, today.month, today.year),
        goals=db.get_goals(user),
        monthly_exp=sum(r[2] for r in rows
                        if str(r[4]).startswith(f"{today.year}-{today.month:02d}")),
        this_month_total=sum(r[2] for r in rows
                        if str(r[4]).startswith(f"{today.year}-{today.month:02d}")),
        monthly_inc=0,
        inc_total=0,
        total_pages=1,
        page_num=1,
        stats=db.get_all_stats(user),
    )



# ── Auth guard ────────────────────────────────────────────────────────────────
def login_required(f):
    from functools import wraps
    @wraps(f)
    def decorated(*args, **kwargs):
        if "user" not in session:
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return decorated

def current_user():
    return session.get("user", "")

# ── Auth routes ───────────────────────────────────────────────────────────────
@app.route("/login", methods=["GET", "POST"])
def login():
    error = None
    if request.method == "POST":
        username = request.form.get("username", "").strip()[:64]
        password = request.form.get("password", "")
        if db.verify_user(username, password):
            session["user"] = username
            return redirect(url_for("dashboard"))
        error = "Invalid username or password."
    return render_template("login.html", error=error, success=None, mode="login")

@app.route("/register", methods=["GET", "POST"])
def register():
    error = None
    success = None
    if request.method == "POST":
        fullname = request.form.get("fullname", "").strip()[:100]
        username = request.form.get("username", "").strip()[:64]
        email    = request.form.get("email", "").strip()[:200]
        password = request.form.get("password", "")
        confirm  = request.form.get("confirm", "")
        if not fullname or not username or not password:
            error = "Full name, username and password are required."
        elif len(username) < 3:
            error = "Username must be at least 3 characters."
        elif " " in username:
            error = "Username cannot contain spaces."
        elif len(password) < 6:
            error = "Password must be at least 6 characters."
        elif password != confirm:
            error = "Passwords do not match."
        elif db.user_exists(username):
            error = f"Username '{username}' is already taken."
        else:
            db.register_user(username, password, fullname, email)
            success = f"Account created! Welcome, {fullname}. Please sign in."
            return render_template("login.html", error=None,
                                   success=success, mode="login")
    return render_template("login.html", error=error,
                           success=success, mode="register")

@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))

@app.route("/change-password", methods=["POST"])
@login_required
def change_password():
    current = request.form.get("current_password", "")
    new_pw  = request.form.get("new_password", "")
    confirm = request.form.get("confirm_password", "")
    if not db.verify_user(current_user(), current):
        return jsonify({"ok": False, "msg": "Current password is incorrect."})
    if new_pw != confirm:
        return jsonify({"ok": False, "msg": "Passwords do not match."})
    if len(new_pw) < 6:
        return jsonify({"ok": False, "msg": "Password must be at least 6 characters."})
    db.change_password(current_user(), new_pw)
    return jsonify({"ok": True, "msg": "Password changed successfully."})

# ── Main routes ───────────────────────────────────────────────────────────────
@app.route("/")
@login_required
def dashboard():
    user  = current_user()
    rows  = db.view_expenses_records(user)
    stats = db.get_all_stats(user)
    today = date.today()
    monthly = sum(r[2] for r in rows
                  if str(r[4]).startswith(f"{today.year}-{today.month:02d}"))
    recent = rows[:8]
    return render_template("index.html",
                           page="dashboard", recent=recent,
                           today=date.today(), cats=CATS, user=user,
                           **base_ctx(user))

@app.route("/add", methods=["GET", "POST"])
@login_required
def add():
    msg = error = None
    if request.method == "POST":
        title    = request.form.get("title", "").strip()[:100]
        amount_s = request.form.get("amount", "").strip()
        category = request.form.get("category", "")
        date_s   = request.form.get("date", "")
        # Description is optional (the UI says so) — fall back to the
        # category name so the record still has a sensible label.
        if not title:
            title = category or "Expense"
        try:
            amount = float(amount_s)
            if amount <= 0:
                raise ValueError("Amount must be positive")
            datetime.strptime(date_s, "%Y-%m-%d")
            db.add_expense_record(current_user(), title, amount, category, date_s)
            msg = f"Added {amount:,.2f} — {title}"
        except ValueError:
            error = "Enter a valid positive amount and a date (YYYY-MM-DD)."
    u = current_user()
    return render_template("index.html", page="add", cats=CATS,
                           msg=msg, error=error,
                           today=date.today(), user=u, **base_ctx(u))

@app.route("/expenses")
@login_required
def expenses():
    q    = request.args.get("q", "").lower()[:100]
    rows = db.view_expenses_records(current_user())
    if q:
        rows = [r for r in rows
                if q in r[1].lower() or q in r[3].lower() or q in str(r[4])]
    u = current_user()
    ctx = base_ctx(u)
    ctx["total_pages"] = 1
    ctx["page_num"] = 1
    ctx["expenses"] = rows  # override with the search-filtered rows for this page
    return render_template("index.html", page="expenses",
                           rows=rows, q=q, cats=CATS,
                           today=date.today(), user=u, **ctx)

@app.route("/delete/<int:eid>", methods=["POST"])
@login_required
def delete(eid):
    db.delete_expense_by_id(current_user(), eid)
    return redirect(url_for("expenses"))

@app.route("/update/<int:eid>", methods=["POST"])
@login_required
def update(eid):
    user  = current_user()
    amt_s  = request.form.get("amount", "").strip()
    title  = request.form.get("title", "").strip()[:100]
    cat    = request.form.get("category", "").strip()
    date_s = request.form.get("date", "").strip()
    if amt_s:
        try:
            db.update_expense_amount(user, eid, float(amt_s))
        except ValueError:
            pass  # intentional — invalid float, skip silently
    if title:
        db.update_expense_title(user, eid, title)
    if cat and cat != "(keep)":
        db.update_expense_category(user, eid, cat)
    if date_s:
        try:
            datetime.strptime(date_s, "%Y-%m-%d")
            db.update_expense_date(user, eid, date_s)
        except ValueError:
            pass  # intentional — invalid date, skip silently
    return redirect(url_for("expenses"))

@app.route("/report")
@login_required
def report():
    today = date.today()
    m  = int(request.args.get("month", today.month))
    yr = int(request.args.get("year",  today.year))
    rows  = db.get_expenses_for_month(current_user(), m, yr)
    total = sum(r[2] for r in rows)
    cat_t = {}
    for r in rows:
        cat_t[r[3]] = cat_t.get(r[3], 0) + r[2]
    cat_data   = sorted(cat_t.items(), key=lambda x: -x[1])
    month_name = datetime(2000, m, 1).strftime("%B")
    u = current_user()
    inc_rows  = db.get_income_for_month(u, m, yr) if hasattr(db, 'get_income_for_month') else []
    ctx = base_ctx(u)
    ctx["inc_total"] = sum(r[2] for r in inc_rows)
    return render_template("index.html", page="report",
                           rows=rows, total=total, cat_data=cat_data,
                           month=m, year=yr, month_name=month_name,
                           cats=CATS, today=today, user=u, **ctx)

# ── Export routes ─────────────────────────────────────────────────────────────
@app.route("/export/excel")
@login_required
def export_excel():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return "Install openpyxl: pip install openpyxl", 500

    rows = db.view_expenses_records(current_user())
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Expenses"
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    hdr_fill = PatternFill("solid", fgColor="1a2035")
    headers  = ["ID", "Title", "Amount", "Category", "Date"]
    col_w    = [8, 35, 18, 18, 15]
    for i, (h, w) in enumerate(zip(headers, col_w), 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font      = hdr_font
        c.fill      = hdr_fill
        c.alignment = Alignment(horizontal="center")
        ws.column_dimensions[c.column_letter].width = w
    for ri, row in enumerate(rows, 2):
        for ci, val in enumerate(row, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.alignment = Alignment(horizontal="center" if ci != 2 else "left")
        ws.cell(row=ri, column=3).number_format = "#,##0.00"
    last = len(rows) + 2
    ws.cell(row=last, column=2, value="TOTAL").font = Font(bold=True)
    ws.cell(row=last, column=3,
            value=sum(r[2] for r in rows)).font = Font(bold=True)
    ws.cell(row=last, column=3).number_format = "#,##0.00"
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name=f"expenses_{date.today()}.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.route("/export/pdf")
@login_required
def export_pdf():
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.platypus import (SimpleDocTemplate, Table,
                                        TableStyle, Paragraph, Spacer)
        from reportlab.lib.styles import getSampleStyleSheet
    except ImportError:
        return "Install reportlab: pip install reportlab", 500

    rows  = db.view_expenses_records(current_user())
    total = sum(r[2] for r in rows)
    buf   = io.BytesIO()
    doc   = SimpleDocTemplate(buf, pagesize=A4,
                              leftMargin=40, rightMargin=40,
                              topMargin=40, bottomMargin=40)
    styles = getSampleStyleSheet()
    story  = [
        Paragraph("Expense Ledger Report", styles["Title"]),
        Paragraph(f"Generated: {date.today().strftime('%d %B %Y')} | User: {current_user()}",
                  styles["Normal"]),
        Spacer(1, 16),
    ]
    data = [["ID", "Title", "Amount", "Category", "Date"]]
    for r in rows:
        data.append([str(r[0]), r[1], f"{r[2]:,.2f}", r[3], str(r[4])])
    data.append(["", "TOTAL", f"{total:,.2f}", "", ""])
    t = Table(data, colWidths=[35, 200, 90, 90, 80])
    t.setStyle(TableStyle([
        ("BACKGROUND",   (0,0), (-1,0),  colors.HexColor("#1a2035")),
        ("TEXTCOLOR",    (0,0), (-1,0),  colors.white),
        ("FONTNAME",     (0,0), (-1,0),  "Helvetica-Bold"),
        ("ALIGN",        (0,0), (-1,-1), "CENTER"),
        ("ALIGN",        (1,1), (1,-1),  "LEFT"),
        ("ROWBACKGROUNDS",(0,1),(-1,-2), [colors.HexColor("#f7f9ff"), colors.white]),
        ("FONTNAME",     (0,-1),(-1,-1), "Helvetica-Bold"),
        ("BACKGROUND",   (0,-1),(-1,-1), colors.HexColor("#e8eaf6")),
        ("GRID",         (0,0), (-1,-1), 0.4, colors.HexColor("#c8d0e8")),
        ("TOPPADDING",   (0,0), (-1,-1), 6),
        ("BOTTOMPADDING",(0,0), (-1,-1), 6),
    ]))
    story.append(t)
    doc.build(story)
    buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name=f"expenses_{date.today()}.pdf",
                     mimetype="application/pdf")

# ── REST API ──────────────────────────────────────────────────────────────────
@app.route("/api/expenses", methods=["GET"])
@login_required
@csrf.exempt
def api_get_expenses():
    """GET /api/expenses — list all expenses for the logged-in user.
    Optional query params: ?category=Food  ?month=5&year=2026
    """
    user     = current_user()
    category = request.args.get("category")
    month    = request.args.get("month",  type=int)
    year     = request.args.get("year",   type=int)

    if month and year:
        rows = db.get_expenses_for_month(user, month, year)
    else:
        rows = db.view_expenses_records(user)

    if category:
        rows = [r for r in rows if r[3].lower() == category.lower()]

    return jsonify([{
        "id":       r[0],
        "title":    r[1],
        "amount":   r[2],
        "category": r[3],
        "date":     r[4],
    } for r in rows])

@app.route("/api/expenses", methods=["POST"])
@login_required
@csrf.exempt
def api_add_expense():
    """POST /api/expenses — add a new expense.
    Body JSON: { title, amount, category, date }
    """
    data = request.get_json(silent=True)
    if not data:
        return jsonify({"error": "JSON body required"}), 400

    title    = str(data.get("title",    "")).strip()[:100]
    category = str(data.get("category", "Other")).strip()
    date_s   = str(data.get("date",     date.today().isoformat()))

    if not title:
        return jsonify({"error": "title is required"}), 400

    try:
        amount = float(data.get("amount", 0))
        if amount <= 0:
            raise ValueError("amount must be positive")
        datetime.strptime(date_s, "%Y-%m-%d")
    except ValueError as e:
        return jsonify({"error": str(e)}), 400

    db.add_expense_record(current_user(), title, amount, category, date_s)
    return jsonify({"ok": True, "message": "Expense added"}), 201

@app.route("/api/expenses/<int:eid>", methods=["DELETE"])
@login_required
@csrf.exempt
def api_delete_expense(eid):
    """DELETE /api/expenses/<id> — delete an expense by ID."""
    db.delete_expense_by_id(current_user(), eid)
    return jsonify({"ok": True, "message": f"Expense {eid} deleted"})

@app.route("/api/expenses/<int:eid>", methods=["PATCH"])
@login_required
@csrf.exempt
def api_update_expense(eid):
    """PATCH /api/expenses/<id> — update title, amount or category."""
    data = request.get_json(silent=True)
    if not data:
        return jsonify({"error": "JSON body required"}), 400
    user = current_user()
    if "amount" in data:
        try:
            db.update_expense_amount(user, eid, float(data["amount"]))
        except ValueError:
            return jsonify({"error": "Invalid amount"}), 400
    if "title" in data:
        db.update_expense_title(user, eid, str(data["title"])[:100])
    if "category" in data:
        db.update_expense_category(user, eid, str(data["category"]))
    return jsonify({"ok": True, "message": "Expense updated"})

@app.route("/api/stats", methods=["GET"])
@login_required
@csrf.exempt
def api_stats():
    """GET /api/stats — summary stats for the logged-in user."""
    return jsonify(db.get_all_stats(current_user()))

@app.route("/api/expenses/summary", methods=["GET"])
@login_required
@csrf.exempt
def api_summary():
    """GET /api/expenses/summary?month=5&year=2026 — per-category breakdown."""
    today = date.today()
    month = request.args.get("month", today.month, type=int)
    year  = request.args.get("year",  today.year,  type=int)
    rows  = db.get_expenses_for_month(current_user(), month, year)
    cat_totals = {}
    for r in rows:
        cat_totals[r[3]] = cat_totals.get(r[3], 0) + r[2]
    return jsonify({
        "month": month, "year": year,
        "total": sum(r[2] for r in rows),
        "by_category": cat_totals,
    })

# ── Error handlers ────────────────────────────────────────────────────────────
# Previously unregistered: users saw Flask/Werkzeug's raw error pages
# (including the bare "Bad Request: The CSRF token is missing." text)
# instead of the templates that already existed in templates/.
from flask_wtf.csrf import CSRFError

@app.errorhandler(CSRFError)
def handle_csrf_error(e):
    return render_template(
        "login.html", mode="login",
        error="Your session has expired. Please refresh the page and try again.",
        success=None
    ), 400

@app.errorhandler(404)
def handle_404(e):
    return render_template("404.html"), 404

@app.errorhandler(500)
def handle_500(e):
    return render_template("500.html"), 500

if __name__ == "__main__":
    app.run(debug=False, port=5000)
